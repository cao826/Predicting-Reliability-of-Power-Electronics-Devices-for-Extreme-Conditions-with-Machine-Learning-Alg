import re
import glob
import pandas as pd
import numpy as np
import math
from decimal import *
import operator
import matplotlib.pyplot as plt
import seaborn as sn
from sklearn import metrics
from sklearn import preprocessing
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import classification_report
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.ensemble import GradientBoostingClassifier
# from xgboost import XGBClassifier
from sklearn.preprocessing import MinMaxScaler
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
from sklearn.utils import resample
from sklearn.cluster import AgglomerativeClustering
import scipy.cluster.hierarchy as shc
# from imblearn.over_sampling import SMOTE
import seaborn as sns
import openpyxl as pxl
from openpyxl import load_workbook
import os.path
import xlsxwriter
from xlsxwriter import Workbook


##########################################################################################
#                Obtaining Device Paths and device_pipeline.xlsx Data                    #
##########################################################################################



# Stores device and device_pipeline.xlsx paths
base_path = '/Users/calchuchesta/Box/Prime technical folder ML and AI work/Jade\'s work/End folder/Processed_Data_Sorted/'
device_paths  = [path.replace('\\', '/') for path in glob.glob(base_path+'transformed_data/*/FIT*/*/*.csv')]
device_pipeline = base_path+'device_pipeline.xlsx'

# Reads data from Dynamic Device Overview, One Hot Encoded, Pre-Static Drain-Source and Pre-Static Gate-Source
# sheets in device_pipline.xlsx
overview = pd.read_excel(device_pipeline, sheet_name='Dynamic Devices Overview')
one_hot_encoded = pd.read_excel(device_pipeline, sheet_name='One Hot Encoded')
prestatic_ds = pd.read_excel(device_pipeline, sheet_name='Pre-Static Drain-Source')
prestatic_gs = pd.read_excel(device_pipeline, sheet_name='Pre-Static Gate-Source')

# Turns the excel sheets into dataframes 
df_pre_static_ds = pd.DataFrame(prestatic_ds[prestatic_ds.columns[:501]])
df_pre_static_gs = pd.DataFrame(prestatic_gs[prestatic_gs.columns[1:501]])
df_one_hot_encoded = pd.DataFrame(one_hot_encoded[one_hot_encoded.columns[5:27]])

# Creates dataframe that will hold the combined 'I at # V' values
df_pre_static_all = df_pre_static_ds.copy()

# Populates df_pre_static_all with df_pre_static_gs data
for column, values in df_pre_static_gs.iteritems():
    # Adds '_GS' to column names to differentiate between Drain-Source and Gate-Source data
    df_pre_static_all[column+'_GS'] = values

# Adds Temperature and Corected Bias Voltage columns from Dynamic Devices Overview
df_pre_static_all['Temperature'] = overview['Temperature']
df_pre_static_all['Corrected Bias Voltage'] = overview['Corrected Bias Voltage']

# Adds one-hot encoded items
df_pre_static_all = pd.concat([df_pre_static_all, df_one_hot_encoded], axis=1, sort=False)

# Adds an empty column called 'Pass/Fail'
df_pre_static_all['Pass/Fail'] = ''

# Dataframe that holds all devices except for TR650V and TR900V devices and Infineon1200V_1
df_pre_static_no_tr = df_pre_static_all.copy()
df_pre_static_no_tr.drop(df_pre_static_no_tr[df_pre_static_no_tr['Device ID'].str.startswith('TR')].index, inplace=True)
df_pre_static_no_tr.drop(df_pre_static_no_tr[df_pre_static_no_tr['Device ID'] == 'Infineon1200V_1'].index, inplace=True)
df_pre_static_no_tr.reset_index(drop=True, inplace=True)

# Drops Infineon1200V_1
df_pre_static = df_pre_static_all.copy()
df_pre_static.drop(df_pre_static[df_pre_static['Device ID'] == 'Infineon1200V_1'].index, inplace=True)
df_pre_static.reset_index(drop=True, inplace=True)





#############################################################################################

# Populates the Pass/Fail column with 'failed' or 'good' depending on each device's path in transformed_data
def passfail_populator(df: pd.DataFrame, paths: list):
    dict_pf = {}
    df_device_id = df['Device ID']

    # Parses device_paths and creates a dictionary matching whether or not a device passed or failed
    for index in range(len(paths)):
        parsed_manufacturer = paths[index].split('/')[-4]
        parsed_num          = re.sub('[^\d]','', paths[index].split('/')[-1])
        man_num             = parsed_manufacturer+'_'+parsed_num
        pass_fail           = paths[index].split('/')[-2]

        dict_pf[man_num] = pass_fail

    # Populates the 'Pass/Fail' column
    for device_id, pf in dict_pf.items():
        index = df_device_id[df_device_id == device_id].index[0]
        df.at[index,'Pass/Fail'] = pf

passfail_populator(df_pre_static_all, device_paths)

##########################################################################################
#                                      ML Functions                                      #
##########################################################################################

# Populates the (empty) test dataframes with test devices, and removes those test devices from df_train
# Each test dataframe consists of device numbers that correspond with model_num
# Ex: If model_num=1, CREE1200V_1, Infineon1200V_1, etc. would be in the test dataframe
def df_test_creator(df_train: pd.DataFrame, df_test: pd.DataFrame, model_num: str):

    # Creates an empty list that will hold the device numbers of test devices
    dev_num_list = []

    # Populates dev_num_list
    for column in df_train['Device ID']:
        dev_num = column.split('_', 1)[1]
        dev_num_list.append(dev_num)

    # Converts dev_num_list to a series
    dev_num_ser = pd.Series((v for v in dev_num_list))

    # Populates df_test and removes devices in this dataframe from df_train
    cond = dev_num_ser == model_num
    rows = df_train.loc[cond, :]
    df_test_new = df_test.append(rows, ignore_index=True)
    df_train_new = df_train.copy()
    df_train_new.drop(rows.index, inplace=True)

    # Resets the index of the training devices
    df_train_new.reset_index(drop=True, inplace=True)

    return(df_test_new, df_train_new)


# Gets confusion matrix values when predicting pass/fail
def get_confusion_matrix_values(y_test, y_pred):
    TP = 0
    FP = 0
    TN = 0
    FN = 0

    for i in range(len(y_pred)):
        if y_test[i]==y_pred[i]=='failed':
           TP += 1
        if y_pred[i]=='failed' and y_test[i]!=y_pred[i]:
           FN += 1
        if y_test[i]==y_pred[i]=='good':
           TN += 1
        if y_pred[i]=='good' and y_test[i]!=y_pred[i]:
           FP += 1

    return(TP, FP, FN, TN)

# Gets accuracy, precision, recall, and f-scores and stores them in separate variables
def get_classification_report(y_test, y_pred, avg: str):
    accuracy = metrics.accuracy_score(y_test, y_pred)

    precision_good   = metrics.precision_score(y_test, y_pred, pos_label='good', average=avg)
    recall_good      = metrics.recall_score(y_test, y_pred, pos_label='good', average=avg)
    fscore_good      = metrics.f1_score(y_test, y_pred, pos_label='good', average=avg)

    precision_failed = metrics.precision_score(y_test, y_pred, pos_label='failed', average=avg)
    recall_failed    = metrics.recall_score(y_test, y_pred, pos_label='failed', average=avg)
    fscore_failed    = metrics.f1_score(y_test, y_pred, pos_label='failed', average=avg)

    precision_avg    = metrics.precision_score(y_test, y_pred, average='weighted')
    recall_avg       = metrics.recall_score(y_test, y_pred, average='weighted')
    fscore_avg       = metrics.f1_score(y_test, y_pred, average='weighted')

    return(accuracy, precision_good, recall_good, fscore_good, precision_failed, recall_failed, fscore_failed,
           precision_avg, recall_avg, fscore_avg)

# Data balancing via oversampling
def oversampler(df_x_train: pd.DataFrame, df_y_train: pd.DataFrame):
    # Balance data with oversampling
    X = pd.concat([df_x_train, df_y_train], axis=1)
    
    # Separates minority and majority classes
    passed = X[X['Pass/Fail']=='good']
    failed = X[X['Pass/Fail']=='failed']
    
    # Upsamples minority
    passed_upsampled = resample(passed,
                          replace=True, # sample with replacement
                          n_samples=len(failed), # match number in majority class
                          random_state=27) # reproducible results
    
    # Combines majority and upsampled minority
    upsampled = pd.concat([failed, passed_upsampled])

    new_x_train = upsampled.drop('Pass/Fail', axis=1)
    new_y_train = upsampled['Pass/Fail']
    
    return(new_x_train, new_y_train)

# Data balancing via undersampling
def undersampler(df_x_train: pd.DataFrame, df_y_train: pd.DataFrame):
    X = pd.concat([df_x_train, df_y_train], axis=1)

    # Separates minority and majority classes
    passed = X[X['Pass/Fail']=='good']
    failed = X[X['Pass/Fail']=='failed']

    # Downsample majority
    failed_downsampled = resample(failed,
                                replace = False, # Samples without replacement
                                n_samples = len(passed), # Matches minority n
                                random_state = 27) # Reproducible results

    # Combines minority and downsampled majority
    downsampled = pd.concat([failed_downsampled, passed])

    new_x_train = downsampled.drop('Pass/Fail', axis=1)
    new_y_train = downsampled['Pass/Fail']

    return(new_x_train, new_y_train)

# Data balancing via smote (i.e., generates data)
def smote_balancer(df_x_train: pd.DataFrame, df_y_train: pd.DataFrame):
    sm = SMOTE(random_state=27, sampling_strategy=1.0)
    new_x_train, new_y_train = sm.fit_sample(df_x_train, df_y_train)

    return(new_x_train, new_y_train)

# Converts values to scientific notation if value is not 0
def scientific_notation(number):
    sci_not = "{:.2e}".format(number)

    if sci_not=='0.00e+00':
        return 0
    else:
        return sci_not

# Gets the feature importances and returns a dictionary matching the feature name and score
def importance_identifier(model, alg):
    # Get feature importance for RFC and GBC
    if alg=='RFC' or alg=='GBC':
        importance = model.feature_importances_
    # Gets feature importance for LR
    elif alg=='LR':
        importance = model.coef_[0]

    # Initialize a dataframe to hold
    dict_importance = {}

    # Summarize feature importance
    for feature, score in enumerate(importance):
#         dict_importance['Feature '+str(feature)+' Score'] = scientific_notation(score)
        dict_importance['Feature '+str(feature)+' Score'] = score

    # Plot feature importance
#     plt.bar([x for x in range(len(importance))], importance)
#     plt.show()

    return dict_importance

# Counts the occurence of passed and failed devices when predicting pass/fail
def occurence_counter_pf(df: pd.DataFrame):
    good_count = 0
    failed_count = 0

    for i in df['Pass/Fail']:
        if i=='good':
            good_count+=1
        else:
            failed_count+=1

    return(good_count, failed_count)

# Performs Logistic Regression, Random Forest Classification, or Gradient Boosting Classification algorithm
# on Pass/Fail or Manufacturer and Voltage
# Most of the returned values are used exclusivly when predicting Pass/Fail
def classifier(df_train: pd.DataFrame, df_test: pd.DataFrame, balancer: str, alg: str, plot: bool, model_num: int,
              to_classify: str):
    # Assigns testing and training data
    if to_classify=='Pass/Fail':
        x_train = df_train.drop(['Device ID', 'Pass/Fail'], axis=1)
        x_test = df_test.drop(['Device ID', 'Pass/Fail'], axis=1)
        y_train = df_train['Pass/Fail']
        y_test = df_test['Pass/Fail']
    elif to_classify=='Manufacturer and Voltage':
        x_train = df_train.drop(['Device ID', 'Pass/Fail', 'Manufacturer and Voltage'], axis=1)
        x_test = df_test.drop(['Device ID', 'Pass/Fail', 'Manufacturer and Voltage'], axis=1)
        y_train = df_train['Manufacturer and Voltage']
        y_test = df_test['Manufacturer and Voltage']

#     sc = StandardScaler()
#     x_train = sc.fit_transform(x_train)
#     x_test =sc.fit_transform(x_test)

#     dict_importance = None

    # Selects a method of balancing the data
    if balancer=='OS':
        # Balance data with oversampling
        x_train, y_train = oversampler(x_train, y_train)
    elif balancer=='US':
        # Balance data with undersampling
        x_train, y_train = undersampler(x_train, y_train)
    elif balancer=='SM':
        # Balance data with SMOTE
        x_train, y_train = smote_balancer(x_train, y_train)

    # Selects an algorithm to run
    if alg=='LR':
        # Creates the logistic regression
        logistic_regression = LogisticRegression(max_iter=5000)
        logistic_regression.fit(x_train, y_train)
        y_pred = logistic_regression.predict(x_test)

        # Gets feature importance
        dict_importance = importance_identifier(logistic_regression, alg)
    elif alg=='RFC':
        # Creates random forest classification
        rfc = RandomForestClassifier(n_estimators=100).fit(x_train, y_train)
        y_pred = rfc.predict(x_test)

        # Gets feature importance
        dict_importance = importance_identifier(rfc, alg)
    elif alg=='GBC':
        gbc = GradientBoostingClassifier(n_estimators=160, learning_rate=.1, max_features='sqrt', max_depth=2,
                                            random_state=None)
        gbc.fit(x_train, y_train)
        y_pred = gbc.predict(x_test)

        # Gets feature importance
        dict_importance = importance_identifier(gbc, alg)

#         xgbc = XGBClassifier()
#         xgbc.fit(x_train, y_train)
#         y_pred = xgbc.predict(x_test)


    # Gets confusion matrix
    confusion_matrix = pd.crosstab(y_test, y_pred, rownames=['Actual'], colnames=['Predicted'])
    tp, fp, fn, tn = get_confusion_matrix_values(y_test, y_pred)

    # Get classification report values
    if to_classify=='Pass/Fail':
        (accuracy, precision_good, recall_good, fscore_good, precision_failed, recall_failed, fscore_failed,
               precision_avg, recall_avg, fscore_avg) = get_classification_report(y_test, y_pred, 'binary')
    elif to_classify=='Manufacturer and Voltage':
        (accuracy, precision_good, recall_good, fscore_good, precision_failed, recall_failed, fscore_failed,
               precision_avg, recall_avg, fscore_avg) = get_classification_report(y_test, y_pred, 'micro')

    new_df_train = x_train.copy()
    new_df_train['Pass/Fail'] = y_train

    if plot==True:
        sn.heatmap(confusion_matrix, annot=True)
        print('---------- Model '+str(model_num)+' ----------')

        # Prints accuracy and confusion matrix
        print('Accuracy: ', accuracy)
        plt.show()

        print('Classification Report')
        print(metrics.classification_report(y_test,y_pred))

    return(accuracy, precision_good, recall_good, fscore_good, precision_failed, recall_failed, fscore_failed,
           precision_avg, recall_avg, fscore_avg, tp, fp, fn, tn, new_df_train, dict_importance, y_pred)

# Returns a spreadsheet summarizing the results of the specified algorithm and feature importances for 24 folds/models and 
# the feature importance values. Folds are differentiated by device number (e.g., testing group consists of
# device 1's across all manufacturer for fold/model 1, with the remaining devices as training).
def model_iterator(df_main: pd.DataFrame, balancer: str, alg: str, plot: bool, output: bool, sheet_name: str):
    # Initializes the model number
    model_num = 1
    
    # Creates an empty dictionary that will map the model number to its accuracy
    dict_model_acc = {}
    
    df_importance = pd.DataFrame()

    # Creates empty lists to be added to LogReg_PassFail in device_pipeline
    list_good_count_train = []
    list_failed_count_train = []
    list_good_count_test = []
    list_failed_count_test = []

    list_tp = []
    list_fp = []
    list_fn = []
    list_tn = []

    list_accuracy = []
    list_precision_good = []
    list_recall_good = []
    list_fscore_good = []
    list_precision_failed = []
    list_recall_failed = []
    list_fscore_failed = []
    list_precision_avg = []
    list_recall_avg = []
    list_fscore_avg = []

    # Iterates through each model (1-24) to create a summary for each
    while (model_num<25):
        df_test_initial = pd.DataFrame()
        df_train = df_main.copy()

        # Gathers the test data and removes devices in df_test from df_train
        df_test, df_train = df_test_creator(df_train, df_test_initial, str(model_num))

        # Performs logistic regression
        (accuracy, precision_good, recall_good, fscore_good, precision_failed, recall_failed, fscore_failed, 
               precision_avg, recall_avg, fscore_avg, tp, fp, fn, tn, df_train, 
                dict_importance, y_pred) = classifier(df_train, df_test, balancer, alg, plot, model_num, 'Pass/Fail')
        
        # Creates the data for feature importance 
        if model_num==1:
            df_importance = df_importance.from_dict(dict_importance, orient='index', columns=['Model '+str(model_num)])
        elif model_num>1:
            list_importance = list(dict_importance.values())
            df_importance['Model '+str(model_num)] = list_importance
        
        good_count_train, failed_count_train = occurence_counter_pf(df_train)
        good_count_test, failed_count_test = occurence_counter_pf(df_test)
        
        # Creates a dictionary matching the model_num to its accuracy
        dict_model_acc['Model '+str(model_num)] = accuracy

        # Appends amount of devices passed/failed to dataframe
        list_good_count_train.append(good_count_train)
        list_failed_count_train.append(failed_count_train)
        list_good_count_test.append(good_count_test)
        list_failed_count_test.append(failed_count_test)

        # Appends Confusion Matrix valus to lists
        list_tp.append(tp)
        list_fp.append(fp)
        list_fn.append(fn)
        list_tn.append(tn)

        # Appends Classification Report values to lists
        list_accuracy.append(accuracy)
        list_precision_good.append(precision_good)
        list_recall_good.append(recall_good)
        list_fscore_good.append(fscore_good)
        list_precision_failed.append(precision_failed)
        list_recall_failed.append(recall_failed)
        list_fscore_failed.append(fscore_failed)
        list_precision_avg.append(precision_avg)
        list_recall_avg.append(recall_avg)
        list_fscore_avg.append(fscore_avg)

        model_num += 1
#         break

    df_summary = pd.DataFrame.from_dict(dict_model_acc, orient='index', columns=['Accuracy'])

    # Adds to new dataframe
    df_summary['Amount Passed (Training)'] = list_good_count_train
    df_summary['Amount Failed (Training)'] = list_failed_count_train
    df_summary['Amount Passed (Testing)'] = list_good_count_test
    df_summary['Amount Failed (Testing)'] = list_failed_count_test

    df_summary['True Positive'] = list_tp
    df_summary['False Positive'] = list_fp
    df_summary['False Negative'] = list_fn
    df_summary['True Negative'] = list_tn

    df_summary['Precision (Pass)'] = list_precision_good
    df_summary['Recall (Pass)'] = list_recall_good
    df_summary['F1 Score (Pass)'] = list_fscore_good
    df_summary['Precision (Failed)'] = list_precision_failed
    df_summary['Recall (Failed)'] = list_recall_failed
    df_summary['F1 Score (Failed)'] = list_fscore_failed
    df_summary['Precision (Avg)'] = list_precision_avg
    df_summary['Recall (Avg)'] = list_recall_avg
    df_summary['F1 Score (Avg)'] = list_fscore_avg
    
    df_summary['Algorithm'] = alg
    
    # Plots accuracy bar chart
#     df_summary.plot.bar(y='Accuracy', legend=False, figsize=(8,8))
    print('Avg. Accuracy: '+str(df_summary['Accuracy'].mean()))
        
    # Transpose df_importance and change the column headers
    df_importance = df_importance.T
    list_df_train_headers = list(df_train.drop(['Pass/Fail'], axis=1))
    df_importance.columns = list_df_train_headers
    
    list_top10_df, df_feature_count = find_top10(df_importance)
    
    if output==True:
        sheet_creator(base_path+'passfail_output.xlsx', sheet_name, df_summary, True)
        sheet_creator(base_path+'passfail_output.xlsx', sheet_name+'_Importances', df_importance, True)
        multiple_dfs(list_top10_df, True, df_feature_count, sheet_name+'_ImportancesTop10', 
                     base_path+'passfail_output.xlsx', 1)
    
    return(df_summary, df_importance, list_top10_df, df_feature_count)

# Returns a dataframe that holds the average accuracies of each manufacturer+voltage or voltage for each ML algorithm used
# Used with df_preds in prediction_generator, step1_predict_manu_volt_dev, step2_predict_passfail,
# and step2_predict_passfail_multi
def avg_manu_volt_accuracy(df_pred: pd.DataFrame, multi_alg: bool, alg: str, target: str):
    df_pred_new = df_pred.copy()
    list_manu_volt = []

    # Appends to the list of manufacturers+voltages or voltages, depending on target
    for device_id in df_pred_new['Device ID']:
        manu_volt = device_id.split('_')[0]
        if target=='MV':
            list_manu_volt.append(manu_volt)
            continue
        elif target=='V':
            volt = re.sub('[\D]','', manu_volt)
            volt = volt + 'V'
            list_manu_volt.append(volt)

#     if target=='MV':
#         # Adds the list of manufacturers as a column to the dataframe
#         df_pred_new['Manufacturer'] = list_manu_volt

    list_of_manu_volt = list(dict.fromkeys(list_manu_volt))

    if multi_alg==True:
        # Creates a list of the columns to iterate through
        list_of_pred = ['LR Prediction', 'RFC Prediction', 'GBC Prediction']

    # Creates a new dataframe to be returned
    df_acc = pd.DataFrame()

    # To be put into df_acc
    list_manu_volt2 = []
    list_alg = []
    list_acc = []

    # Iterates through the manufacturers and appends to the lists
    for manuvolt in list_of_manu_volt:
        df_temp = df_pred_new.copy()
        df_temp.drop(df_temp[~df_temp['Device ID'].str.contains(manuvolt)].index, inplace=True)
        df_temp.reset_index(drop=True, inplace=True)

        list_act = df_temp['Actual'].tolist()

        if multi_alg==True:
            alg = None

            # Gets the accuracy for each algorithm
            for pred in list_of_pred:
                list_pred = df_temp[pred]

                count_match = 0
                for act, pred2 in zip(list_act, list_pred):
                    if act==pred2:
                        count_match+=1

                accuracy = count_match / len(list_act)

                list_manu_volt2.append(manuvolt)
                list_alg.append(pred)
                list_acc.append(accuracy)

            list_act = df_temp['Actual'].tolist()
        elif multi_alg==False:
            # Gets the accuracy for specified algorithm
            count_match = 0
            for act, pred2 in zip(list_act, df_temp['Predicted']):
                if act==pred2:
                    count_match+=1

            accuracy = count_match / len(list_act)

            list_manu_volt2.append(manuvolt)
#                 list_alg.append(pred)
            list_acc.append(accuracy)

    if target=='MV':
        col_name = 'Manufacturer'
    elif target=='V':
        col_name = 'Voltage'

    print(list_manu_volt2)
    # Populates the dataframe
    df_acc[col_name] = list_manu_volt2
    if multi_alg==True:
        df_acc['Model'] = list_alg
    elif multi_alg==False:
        df_acc['Model'] = alg
    df_acc['Accuracy'] = list_acc

    if multi_alg==True:
        list_row_acc = ['Prediction Accuracies', np.nan]
        # Gets the overall accuracy for each algorithm, across all manufacturers
        for pred in list_of_pred:
            list_pred = df_pred_new[pred]

            count_match = 0
            for (act, pred2) in zip(df_pred_new['Actual'], list_pred):
                if act==pred2:
                    count_match+=1
            count_nan = int(df_pred_new[pred].isna().sum())

            accuracy = count_match / (int(len(df_pred_new['Actual']) - count_nan))

            list_row_acc.append(accuracy)

        df_pred_new.loc[len(df_pred_new)] = list_row_acc

    return (df_acc, df_pred_new)

# Returns a dataframe of the prediciton results of different algorithms for each manufacturer for 24 folds/models, and
# a second dataframe where devices are split by manufacturer and their accuracies calculated
# Folds are differentiated by device number (e.g., testing group consists of
# device 1's across all manufacturer for fold/model 1, with the remaining devices as training)

def pf_prediction_generator(df_main: pd.DataFrame, balancer: str, plot: bool, output: bool):
    # Initializes the model number
#     model_num = 1

    list_devid = []
    list_actpred = []
    list_ypred_lr = []
    list_ypred_rfc = []
    list_ypred_gbc = []

    # Creates a list to hold the algorithms to be performed
    list_alg = ['LR', 'RFC', 'GBC']

    # Iterates through the list of algorithms and populates df_preds
    for alg in list_alg:
        # Initializes model number
        model_num = 1

        # Iterates through each model (1-24) to create a summary for each
        while (model_num<25):
            df_test_initial = pd.DataFrame()
            df_train = df_main.copy()

            # Gathers the test data and removes devices in df_test from df_train
            df_test, df_train = df_test_creator(df_train, df_test_initial, str(model_num))

            # Performs LR, RFC, GBC
            (accuracy, precision_good, recall_good, fscore_good, precision_failed, recall_failed, fscore_failed,
                   precision_avg, recall_avg, fscore_avg, tp, fp, fn, tn, df_train,
                    dict_importance, y_pred) = classifier(df_train, df_test, balancer, alg, plot, model_num, 'Pass/Fail')

            if alg=='LR':
                list_devid.extend(df_test['Device ID'].tolist())
                list_actpred.extend(df_test['Pass/Fail'].tolist())
                list_ypred_lr.extend(y_pred.tolist())
            elif alg=='RFC':
#                 list_devid.extend(df_test['Device ID'].tolist())
                list_ypred_rfc.extend(y_pred.tolist())
            elif alg=='GBC':
#                 list_devid.extend(df_test['Device ID'].tolist())
                list_ypred_gbc.extend(y_pred.tolist())

            model_num += 1
#             break

    df_preds = pd.DataFrame(list_devid,columns=['Device ID'])

    df_preds['Actual'] = list_actpred
    df_preds['LR Prediction'] = list_ypred_lr
    df_preds['RFC Prediction'] = list_ypred_rfc
    df_preds['GBC Prediction'] = list_ypred_gbc

    # Gets a df with avg. accuracies of predictions grouped by  manufacturer+voltage
    # Adds avg. accuracy for each ML model (LR, RFC, GBC) to df_preds
    df_preds_acc, df_preds = avg_manu_volt_accuracy(df_preds, True, None, 'MV')

    list_dfs = [df_preds, df_preds_acc]
    if output==True:
        sheet_creator('passfail_output.xlsx', 'PF_Pred_Results', df_preds, True)
        multiple_dfs(list_dfs, False, None, 'PF_Pred_Results', 'passfail_output.xlsx', 1)
    if plot==True:
        boxplotter(df_preds_acc, 'Manufacturer', 'Accuracy', 'Manufacturer', 'Accuracy',
           'Accuracies of Pass/Fail Classification Grouped by Manufacturer', 'BoxPlot_PF_PredGen_Acc', False)

    return (df_preds, df_preds_acc)

# Predict Manufacturer+Voltage or Voltage with train/test sets split 24-fold by device number
# Purpose of this is to weed out (bad) devices that might have been affected by extraneous variables
# Step 1 of 2-Step process
def step1_predict_manu_volt_dev(df: pd.DataFrame, target: str, alg: str, balancer: str, plot: bool, output: bool):
    model_num = 1

    df_main = df.copy()
    df_importance = pd.DataFrame()
    dict_manu_volt = {}

    # Creates lists to hold each device's actual and predicted classification
    list_devid = []
    list_actpred = []
    list_ypred = []

    # Creates an empty dictionary that will map the model number to its accuracy
    dict_model_acc = {}

    # Drops one-hot encoded items
    df_main = df_main[df_main.columns.drop(list(df_main.filter(regex='Volt')))]
    df_main = df_main[df_main.columns.drop(list(df_main.filter(regex='Manu')))]

    # Gets the manufacturer and its voltage
    for device_id in df_main['Device ID']:
        manu_volt = device_id.split('_')[0]
        if target=='MV':
            manu = re.sub('[\dV]','', manu_volt)
            dict_manu_volt[device_id] = manu_volt
        if target=='V':
            volt = re.sub('[\D]','', manu_volt)
            volt = volt + 'V'
            dict_manu_volt[device_id] = volt

    # Adds manufacturer+voltage to the dataframe
    df_main['Manufacturer and Voltage'] = df_main['Device ID'].map(dict_manu_volt)

    while (model_num<25):
        df_test_initial = pd.DataFrame()
        df_train = df_main.copy()

        # Gathers the test data and removes devices in df_test from df_train
        df_test, df_train = df_test_creator(df_train, df_test_initial, str(model_num))

        # Performs logistic regression
        (accuracy, precision_good, recall_good, fscore_good, precision_failed, recall_failed, fscore_failed,
               precision_avg, recall_avg, fscore_avg, tp, fp, fn, tn, df_train,
                dict_importance, y_pred) = classifier(df_train, df_test, balancer, alg, plot, model_num,
                                                      'Manufacturer and Voltage')

        list_devid.extend(df_test['Device ID'].tolist())
        list_actpred.extend(df_test['Manufacturer and Voltage'].tolist())
        list_ypred.extend(y_pred.tolist())

        # Creates the data for feature importance
        if model_num==1:
            df_importance = df_importance.from_dict(dict_importance, orient='index', columns=['Model '+str(model_num)])
        elif model_num>1:
            list_importance = list(dict_importance.values())
            df_importance['Model '+str(model_num)] = list_importance

        # Creates a dictionary matching the model_num to its accuracy
        dict_model_acc['Model '+str(model_num)] = accuracy

        model_num += 1
#         break

    df_summary = pd.DataFrame.from_dict(dict_model_acc, orient='index', columns=['Accuracy'])
    df_summary['Algorithm'] = alg

    # Creates a dataframe with the predictions for each Device ID
    df_preds = pd.DataFrame(list_devid,columns=['Device ID'])
    df_preds['Actual'] = list_actpred
    df_preds['Predicted'] = list_ypred

    # Creates an empty list that will hold whether or not the model correctly predicted the manufacturer
    list_pred_eval = []

    # Checks to see whether the manufacturer was correctly predicted and appends to list
    for a, p in zip(df_preds['Actual'], df_preds['Predicted']):
        if a==p:
            list_pred_eval.append('Correct')
        else:
            list_pred_eval.append('Incorrect')

    # Adds the evaluation list to df_preds
    df_preds['Evaluation'] = list_pred_eval

    # Plots accuracy bar chart
#     df_summary.plot.bar(y='Accuracy', legend=False, figsize=(8,8))
    print('Avg. Accuracy: '+str(df_summary['Accuracy'].mean()))

    # Transpose df_importance and change the column headers
    df_importance = df_importance.T
    list_df_train_headers = list(df_train.drop(['Pass/Fail'], axis=1))
    df_importance.columns = list_df_train_headers

    # Get the accuracy by manufacturer
    df_preds_acc, df_preds = avg_manu_volt_accuracy(df_preds, False, alg, 'MV')

    list_preds = [df_preds, df_preds_acc]
    # Exports to passfail_output.xlsx
    if output==True and target=='V':
        multiple_dfs(list_preds, False, None, 'Step1_PredictVolt', base_path+'test_output.xlsx', 1)
    if output==True and target=='MV':
        multiple_dfs(list_preds, False, None, 'Step1_PredictManuVolt', base_path+'test_output.xlsx', 1)

    return(df_summary, df_importance, df_preds, df_preds_acc)

# Predict Pass/Fail of devices whose manufacturers were correctly predicted
# Takes df_preds  of step1_predict_manu_volt_dev and the dataframe with the features to be used for ML
# Step 2 of 2-Step process
def step2_predict_passfail(df_main: pd.DataFrame, df_pred: pd.DataFrame, target: str, multi_alg: bool, alg: str,
                           balancer: str, plot: bool, output: bool):
    df_features = df_main.copy()
    df_manu_preds = df_pred.copy()

    # Drops one-hot encoded items
    df_features = df_features[df_features.columns.drop(list(df_features.filter(regex='Volt')))]
    df_features = df_features[df_features.columns.drop(list(df_features.filter(regex='Manu')))]

    # Drops rows that are incorrectly evaluated
    df_manu_preds.drop(df_manu_preds[df_manu_preds['Evaluation'] == 'Incorrect'].index, inplace=True)

    # Merges the two dataframe, preserving only correctly classified devices
    df_features = pd.merge(df_features, df_manu_preds, on='Device ID', how='inner')
    df_features.reset_index(drop=True, inplace=True)

    df_features.drop(['Evaluation', 'Predicted', 'Actual'], axis=1, inplace=True)

    # Creates a list that will hold the list of manufacturers
    list_of_manu = []

    for device_id in df_features['Device ID']:
        manu_volt = device_id.split('_')[0]
        list_of_manu.append(manu_volt)

    # Removes duplicates
    list_of_manu = list(dict.fromkeys(list_of_manu))

    # Creates lists to hold each device's actual and predicted classification
    list_devid = []
    list_actpred = []
    list_ypred = []
    list_ypred_lr = []
    list_ypred_rfc = []
    list_ypred_gbc = []

    # Creates a list to hold the algorithms to be performed
    if multi_alg==True:
        list_alg = ['LR', 'RFC', 'GBC']
    elif multi_alg==False:
        list_alg = [alg]

    count_skipped = 0

    # Iterates through the list of algorithms and populates df_preds
    for alg in list_alg:
        # Initializes model number
        model_num = 1

        # Iterates through each manufacturer and performs ML
        for manu in list_of_manu:
            df_temp = df_features.copy()
            df_temp.drop(df_temp[~df_temp['Device ID'].str.startswith(manu)].index, inplace=True)
            df_temp.reset_index(drop=True, inplace=True)

            # Creates a list of the device numbers that will be iterated through
            list_dev_num = []

            for device_id in df_temp['Device ID']:
                dev_num = int(device_id.split('_')[1])
                list_dev_num.append(dev_num)

            model_num = 1
            # Iterates through each device number of the current manufacturer, and uses the current model_num for the test data
            while (model_num<max(list_dev_num)):
                print(str(model_num)+' '+manu+' '+alg)

                # If the device number isn't in the list, skip
                if model_num not in list_dev_num:
                    model_num += 1
                    continue

                df_test_initial = pd.DataFrame()
                df_train = df_temp.copy()

                # Gathers the test data and removes devices in df_test from df_train
                df_test, df_train = df_test_creator(df_train, df_test_initial, str(model_num))

                # Checks to see if the manufacturer has any passed devices
                # If not, returns to the beginning of the loop and skips that manufacturer
                if alg=='LR':
                    if not df_train['Pass/Fail'].str.contains('good').any() or not df_train['Pass/Fail'].str.contains('good').any():
                        print('This training set does not contain both failed and passed devices.')
                        list_devid.extend(df_test['Device ID'].tolist())
                        list_actpred.extend(df_test['Pass/Fail'].tolist())
                        list_ypred_lr.append(np.nan)
                        model_num+=1
#                         count_skipped+=1
                        continue
                if alg=='GBC':
                    if not df_train['Pass/Fail'].str.contains('good').any() or not df_train['Pass/Fail'].str.contains('good').any():
                        print('This training set does not contain both failed and passed devices.')
                        list_ypred_gbc.append(np.nan)
                        model_num+=1
#                         count_skipped+=1
                        continue

                # Performs LR, RFC, GBC
                (accuracy, precision_good, recall_good, fscore_good, precision_failed, recall_failed, fscore_failed,
                       precision_avg, recall_avg, fscore_avg, tp, fp, fn, tn, df_train,
                        dict_importance, y_pred) = classifier(df_train, df_test, balancer, alg, plot, model_num, 'Pass/Fail')

                if alg=='LR':
                    list_devid.extend(df_test['Device ID'].tolist())
                    list_actpred.extend(df_test['Pass/Fail'].tolist())
                    list_ypred_lr.extend(y_pred.tolist())
                elif alg=='RFC':
                    list_ypred_rfc.extend(y_pred.tolist())
                elif alg=='GBC':
                    list_ypred_gbc.extend(y_pred.tolist())

                model_num += 1

    # Creates a dataframe with the predictions for each Device ID
    df_preds = pd.DataFrame(list_devid,columns=['Device ID'])

    df_preds['Actual'] = list_actpred
    if multi_alg==True:
        df_preds['LR Prediction'] = list_ypred_lr
        df_preds['RFC Prediction'] = list_ypred_rfc
        df_preds['GBC Prediction'] = list_ypred_gbc

        df_preds_acc, df_preds = avg_manu_volt_accuracy(df_preds, True, None, target)
    elif multi_alg==False and alg=='LR':
        df_preds['Predicted'] = list_ypred_lr
        df_preds_acc, df_preds = avg_manu_volt_accuracy(df_preds, False, alg, target)
    elif multi_alg==False and alg=='RFC':
        df_preds['Predicted'] = list_ypred_rfc
        df_preds_acc, df_preds = avg_manu_volt_accuracy(df_preds, False, alg, target)
    elif multi_alg==False and alg=='GBC':
        df_preds['Predicted'] = list_ypred_gbc
        df_preds_acc, df_preds = avg_manu_volt_accuracy(df_preds, False, alg, target)

    # Prints Manufacturer/Voltage vs Accuracy box plot
    # Change 'False' to 'True' to save figure
    if plot==True and target=='V':
        boxplotter(df_preds_acc, 'Voltage', 'Accuracy', 'Voltage', 'Accuracy',
           'Accuracies of Pass/Fail Classification Grouped by Voltage', 'BoxPlot_Volt_PF_Acc', True)
    elif plot==True and target=='MV':
        boxplotter(df_preds_acc, 'Manufacturer', 'Accuracy', 'Manufacturer', 'Accuracy',
           'Accuracies of Pass/Fail Classification Grouped by Manufacturer', 'BoxPlot_ManuVolt_PF_Acc', True)

    list_preds = [df_preds, df_preds_acc]
    # Export sheet to passfail_output.xlsx
    if output==True and target=='V':
        multiple_dfs(list_preds, False, None, 'Step2_V_PredictPF', base_path+'passfail_output.xlsx', 1)
    elif output==True and target=='MV':
        multiple_dfs(list_preds, False, None, 'Step2_MV_PredictPF', base_path+'passfail_output.xlsx', 1)


    return (df_preds, df_preds_acc)

# Takes the specified device number of each manufacturer and predicts the manufacturer+voltage
# Prints one confusion matrix and classification report with results
# Ex: test_dev_num=1, all device 1's are in the testing data, with the remaining devices in training
def predict_manuvolt(drop_ohe: bool, alg: str, test_dev_num: int):
    df_predict_manu = df_pre_static_all.copy()
    df_importance = pd.DataFrame()
    dict_manu_volt = {}

    # Drops one-hot encoded items
    if drop_ohe==True:
        df_predict_manu = df_predict_manu[df_predict_manu.columns.drop(list(df_predict_manu.filter(regex='Volt')))]
        df_predict_manu = df_predict_manu[df_predict_manu.columns.drop(list(df_predict_manu.filter(regex='Manu')))]

    # Gets the manufacturer and its voltage
    for device_id in df_predict_manu['Device ID']:
        manu_volt = device_id.split('_')[0]
        manu = re.sub('[\dV]','', manu_volt)
        dict_manu_volt[device_id] = manu_volt

    # Adds the manufacturer+voltage to the dataframe
    df_predict_manu['Manufacturer and Voltage'] = df_predict_manu['Device ID'].map(dict_manu_volt)

    # Holds manufacturer
#     list_manu = ['CREE1200V', 'CREE1700V', 'Infineon1200V', 'LTF1200V', 'LTF1700V', 'ROHM1200V', 'ROHM1700V',
#                'ST1200V', 'TR650V', 'TR900V']
    list_manu = ['C1200V', 'G1700V', 'F1200V', 'D1200V', 'E1700V', 'B1200V', 'J1700V', 'A1200V',
                          'H650', 'I900']

    # Change manufacturer to their codenames
    for m, c in zip(list_manu, list_manu):
        df_predict_manu['Manufacturer and Voltage'] = df_predict_manu['Manufacturer and Voltage'].str.replace(m, c)

    df_test_initial = pd.DataFrame()
    df_train = df_predict_manu.copy()

    # Gathers all device 1's across manufacturers and uses it as test data
    # Removes devices in df_test from df_train
    df_test, df_train = df_test_creator(df_train, df_test_initial, str(test_dev_num))

    x_train = df_train.drop(['Device ID', 'Manufacturer and Voltage', 'Pass/Fail'], axis=1)
    x_test = df_test.drop(['Device ID', 'Manufacturer and Voltage', 'Pass/Fail'], axis=1)
    y_train = df_train['Manufacturer and Voltage']
    y_test = df_test['Manufacturer and Voltage']

#     x = df_predict_manu.drop(['Device ID', 'Pass/Fail', 'Manufacturer and Voltage'], axis=1)
#     y = df_predict_manu['Manufacturer and Voltage']

    # Testing is done on 20% of the data, training on 80%
#     x_train,x_test,y_train,y_test = train_test_split(x,y,test_size=0.1,random_state=0)

    # Performs Logistic Regression
    if alg=='LR':
        logistic_regression= LogisticRegression(max_iter=2100)
        logistic_regression.fit(x_train,y_train)
        y_pred = logistic_regression.predict(x_test)
    # Performs Random Forest Classification
    elif alg=='RFC':
        rfc = RandomForestClassifier(n_estimators=100).fit(x_train, y_train)
        y_pred = rfc.predict(x_test)
    # Performs Gradient Boosting Classification
    elif alg=='GBC':
        gbc = GradientBoostingClassifier(n_estimators=160, learning_rate=.1, max_features='sqrt', max_depth=2,
                                            random_state=None)
        gbc.fit(x_train, y_train)
        y_pred = gbc.predict(x_test)

    # Gets confusion matrix
    confusion_matrix = pd.crosstab(y_test, y_pred, rownames=['Actual'], colnames=['Predicted'])
    sn.heatmap(confusion_matrix, annot=True)

    # Prints accuracy and confusion matrix
    print('Accuracy: ',metrics.accuracy_score(y_test, y_pred))

    print(classification_report(y_test,y_pred))

    plt.show()

# Predicts Pass/Fail, using entire manufacturer sets as the testing set (e.g., all of CREE1200V), and the remaining
# for training (e.g., all but CREE1200V)

def predict_manutest_pf(alg: str):
    # Includes Infineon1200V_1
    df_predict_manu_pf = df_pre_static.copy()
    list_manu_volt = []
    dev_num = 1

    # Creates an empty dictionary that will map the model number to its accuracy
    dict_model_acc = {}

    df_importance = pd.DataFrame()

    # Gets the manufacturer and its voltage
    for device_id in df_pre_static['Device ID']:
        manu_volt = device_id.split('_')[0]
        list_manu_volt.append(manu_volt)
  
    # Removes duplicates from the list
    list_manu_volt = list(dict.fromkeys(list_manu_volt))

    count = 1
    model_num = dev_num
    for manu_volt in list_manu_volt:
        x_test = df_pre_static.copy()
        x_test.drop(x_test[~x_test['Device ID'].str.startswith(manu_volt)].index, inplace=True)
        x_test.reset_index(drop=True, inplace=True)
        y_test = x_test['Pass/Fail']
        x_test = x_test.drop(['Device ID', 'Pass/Fail'], axis=1)
        
        x_train = df_pre_static.copy()
        x_train.drop(x_train[x_train['Device ID'].str.startswith(manu_volt)].index, inplace=True)
        x_train.reset_index(drop=True, inplace=True)
        y_train = x_train['Pass/Fail']
        x_train = x_train.drop(['Device ID', 'Pass/Fail'], axis=1)
        
    # Checks to see if the manufacturer has any passed devices
    # If not, returns to the beginning of the loop and skips that manufacturer
        if alg=='LR' or alg=='GBC':
            if not y_train.str.contains('good').any():
                continue

        # Performs logistic Regression
        if alg=='LR':
            logistic_regression= LogisticRegression(max_iter=2100)
            logistic_regression.fit(x_train,y_train)
            y_pred = logistic_regression.predict(x_test)

            dict_importance = importance_identifier(logistic_regression, alg)
        # Performs Random Forest Classification
        if alg=='RFC':
            rfc = RandomForestClassifier(n_estimators=100).fit(x_train, y_train)
            y_pred = rfc.predict(x_test)

            dict_importance = importance_identifier(rfc, alg)
        # Performs Gradient Boosting Classification
        if alg=='GBC':
            gbc = GradientBoostingClassifier(n_estimators=160, learning_rate=.1, max_features='sqrt', max_depth=2, 
                                        random_state=None)
            gbc.fit(x_train, y_train)
            y_pred = gbc.predict(x_test)
            
            dict_importance = importance_identifier(gbc, alg)
            

        # Gets confusion matrix
        confusion_matrix = pd.crosstab(y_test, y_pred, rownames=['Actual'], colnames=['Predicted'])
        tp, fp, fn, tn = get_confusion_matrix_values(y_test, y_pred)

        # Get classification report values
        (accuracy, precision_good, recall_good, fscore_good, precision_failed, recall_failed, fscore_failed, 
               precision_avg, recall_avg, fscore_avg) = get_classification_report(y_test, y_pred, 'binary')

    #     new_df_train = x_train.copy()
    #     new_df_train['Pass/Fail'] = y_train

        # Gets confusion matrix
        #confusion_matrix = pd.crosstab(y_test, y_pred, rownames=['Actual'], colnames=['Predicted'])
        #sn.heatmap(confusion_matrix, annot=True)

        # Prints accuracy and confusion matrix
        print('----------------'+manu_volt+'----------------')

        print('Accuracy: ',metrics.accuracy_score(y_test, y_pred))
        plt.show()

        print(classification_report(y_test,y_pred))

        # Creates the data for feature importance 
        if count==1:
            df_importance = df_importance.from_dict(dict_importance, orient='index', columns=['Manufacturer '+manu_volt])
        elif count>1:
            list_importance = list(dict_importance.values())
            df_importance['Manufacturer '+manu_volt] = list_importance

        # Creates a dictionary matching the model_num to its accuracy
        dict_model_acc[manu_volt] = accuracy


        count+=1

    #     break

    df_summary = pd.DataFrame.from_dict(dict_model_acc, orient='index', columns=['Accuracy'])
    
    df_summary['Algorithm'] = alg

    # Plots accuracy bar chart
    #     df_summary.plot.bar(y='Accuracy', legend=False, figsize=(8,8))
    print('Avg. Accuracy: '+str(df_summary['Accuracy'].mean()))

    return df_summary

# # Predict Pass/Fail with manufacturers separated
# # Specified dev_num is the test data (1 test device per manufacturer)
# # pd.set_option("display.max_rows", 5, "display.max_columns", 5)
# # pd.set_option("display.max_rows", None, "display.max_columns", None)

def predict_manu_separated_pf(dev_num: str, alg: str):
    # Includes Infineon1200V_1
    df_predict_manu_separated_pf = df_pre_static_all.copy()
    # dict_manu_volt = {}
    list_manu_volt = []

    # Creates an empty dictionary that will map the model number to its accuracy
    dict_model_acc = {}

    df_importance = pd.DataFrame()

    # Creates empty lists to be added to LogReg_PassFail in device_pipeline
    list_good_count_train = []
    list_failed_count_train = []
    list_good_count_test = []
    list_failed_count_test = []

    list_tp = []
    list_fp = []
    list_fn = []
    list_tn = []

    list_accuracy = []
    list_precision_good = []
    list_recall_good = []
    list_fscore_good = []
    list_precision_failed = []
    list_recall_failed = []
    list_fscore_failed = []
    list_precision_avg = []
    list_recall_avg = []
    list_fscore_avg = []

    # Gets the manufacturer and its voltage
    for device_id in df_pre_static['Device ID']:
        manu_volt = device_id.split('_')[0]
        list_manu_volt.append(manu_volt)

    # Adds the manufacturer+voltage to the dataframe
    # df_predict_manu_separated_pf['Manufacturer and Voltage'] = df_predict_manu_separated_pf['Device ID'].map(dict_manu_volt)

    # Removes duplicates from the list
    list_manu_volt = list(dict.fromkeys(list_manu_volt))

    count = 1
    model_num = dev_num
    for manu_volt in list_manu_volt:
        df_single_manu = df_predict_manu_separated_pf.copy()

        df_single_manu.drop(df_single_manu[~df_single_manu['Device ID'].str.startswith(manu_volt)].index, inplace=True)
        df_single_manu.reset_index(drop=True,inplace=True)

        df_test_initial = pd.DataFrame()
        df_train = df_single_manu.copy()

        if alg=='LR' or alg=='GBC':
            if not df_train['Pass/Fail'].str.contains('good').any():
                continue

        # Gathers all device 1's across manufacturers and uses it as test data
        # Removes devices in df_test from df_train
        df_test, df_train = df_test_creator(df_train, df_test_initial, model_num)

        x_train = df_train.drop(['Device ID', 'Pass/Fail'], axis=1)
        x_test = df_test.drop(['Device ID', 'Pass/Fail'], axis=1)
        y_train = df_train['Pass/Fail']
        y_test = df_test['Pass/Fail']

        # Performs logistic Regression
        if alg=='LR':
            logistic_regression= LogisticRegression(max_iter=2100)
            logistic_regression.fit(x_train,y_train)
            y_pred=logistic_regression.predict(x_test)

            dict_importance = importance_identifier(logistic_regression, alg)
        # Performs Random Forest Classification
        if alg=='RFC':
            rfc = RandomForestClassifier(n_estimators=100).fit(x_train, y_train)
            y_pred = rfc.predict(x_test)

            dict_importance = importance_identifier(rfc, alg)
        # Performs Gradient Boosting Classification
        if alg=='GBC':
            gbc = GradientBoostingClassifier(n_estimators=160, learning_rate=.1, max_features='sqrt', max_depth=2,
                                        random_state=None)
            gbc.fit(x_train, y_train)
            y_pred = gbc.predict(x_test)

            dict_importance = importance_identifier(gbc, alg)


        # Gets confusion matrix
        confusion_matrix = pd.crosstab(y_test, y_pred, rownames=['Actual'], colnames=['Predicted'])
        tp, fp, fn, tn = get_confusion_matrix_values(y_test, y_pred)

        # Get classification report values
        (accuracy, precision_good, recall_good, fscore_good, precision_failed, recall_failed, fscore_failed,
               precision_avg, recall_avg, fscore_avg) = get_classification_report(y_test, y_pred, 'binary')

    #     new_df_train = x_train.copy()
    #     new_df_train['Pass/Fail'] = y_train

        # Gets confusion matrix
        confusion_matrix = pd.crosstab(y_test, y_pred, rownames=['Actual'], colnames=['Predicted'])
        sn.heatmap(confusion_matrix, annot=True)

        # Prints accuracy and confusion matrix
        print('----------------'+manu_volt+'----------------')
        print('Accuracy: ',metrics.accuracy_score(y_test, y_pred))
        plt.show()

        print(classification_report(y_test,y_pred))

        # Creates the data for feature importance
        if count==1:
            df_importance = df_importance.from_dict(dict_importance, orient='index', columns=['Manufacturer '+manu_volt])
        elif count>1:
            list_importance = list(dict_importance.values())
            df_importance['Manufacturer '+manu_volt] = list_importance

        good_count_train, failed_count_train = occurence_counter_pf(df_train)
        good_count_test, failed_count_test = occurence_counter_pf(df_test)

        # Creates a dictionary matching the model_num to its accuracy
        dict_model_acc[manu_volt] = accuracy

        # Appends amount of devices passed/failed to dataframe
        list_good_count_train.append(good_count_train)
        list_failed_count_train.append(failed_count_train)
        list_good_count_test.append(good_count_test)
        list_failed_count_test.append(failed_count_test)

        # Appends Confusion Matrix valus to lists
        list_tp.append(tp)
        list_fp.append(fp)
        list_fn.append(fn)
        list_tn.append(tn)

        # Appends Classification Report values to lists
        list_accuracy.append(accuracy)
        list_precision_good.append(precision_good)
        list_recall_good.append(recall_good)
        list_fscore_good.append(fscore_good)
        list_precision_failed.append(precision_failed)
        list_recall_failed.append(recall_failed)
        list_fscore_failed.append(fscore_failed)
        list_precision_avg.append(precision_avg)
        list_recall_avg.append(recall_avg)
        list_fscore_avg.append(fscore_avg)

        count+=1

    df_summary = pd.DataFrame.from_dict(dict_model_acc, orient='index', columns=['Accuracy'])

    # Adds to new dataframe
    df_summary['Amount Passed (Training)'] = list_good_count_train
    df_summary['Amount Failed (Training)'] = list_failed_count_train
    df_summary['Amount Passed (Testing)'] = list_good_count_test
    df_summary['Amount Failed (Testing)'] = list_failed_count_test

    df_summary['True Positive'] = list_tp
    df_summary['False Positive'] = list_fp
    df_summary['False Negative'] = list_fn
    df_summary['True Negative'] = list_tn

    df_summary['Precision (Pass)'] = list_precision_good
    df_summary['Recall (Pass)'] = list_recall_good
    df_summary['F1 Score (Pass)'] = list_fscore_good
    df_summary['Precision (Failed)'] = list_precision_failed
    df_summary['Recall (Failed)'] = list_recall_failed
    df_summary['F1 Score (Failed)'] = list_fscore_failed
    df_summary['Precision (Avg)'] = list_precision_avg
    df_summary['Recall (Avg)'] = list_recall_avg
    df_summary['F1 Score (Avg)'] = list_fscore_avg

    df_summary['Algorithm'] = alg

    # Plots accuracy bar chart
    #     df_summary.plot.bar(y='Accuracy', legend=False, figsize=(8,8))
    print('Avg. Accuracy: '+str(df_summary['Accuracy'].mean()))

    # Transpose df_importance and change the column headers
#     df_importance = df_importance.T
#     list_df_train_headers = list(df_train.drop(['Pass/Fail', 'Device ID'], axis=1))
#     df_importance.columns = list_df_train_headers

#     list_top10_df, df_feature_count = find_top10(df_importance)

#     return(df_summary, df_importance, list_top10_df, df_feature_count)
    return df_summary

# Finds the top 10 features for each model and counts the occurence of each feature
def find_top10(df: pd.DataFrame):
    df_top10_features = pd.DataFrame()
    list_top10_df = []

    # Iterates through the feature importance dataframe and creates a list of the top 10 features of each model
    for model_num, row_values in df.iterrows():
        row_sorted = row_values.sort_values(ascending=False)
        top10 = row_sorted[:10]
        df_top10_features[model_num] = top10
        list_top10_df.append(df_top10_features)
        df_top10_features = pd.DataFrame()

    # Start of the feature count
    list_feature = []

    # Creates a giant list of all of the top 10 features from each model
    for dataframe in list_top10_df:
        list_dataframe = list(dataframe.index)
        list_feature.extend(list_dataframe)

    # Counts the occurence of each feature and converts that into a dictionary (feature: count)
    dict_feature_count = dict((x,list_feature.count(x)) for x in set(list_feature))

    # Sorts dict_feature_count in descending order based off of the values
    dict_feature_count = dict(sorted(dict_feature_count.items(), key=operator.itemgetter(1),reverse=True))
    df_feature_count = pd.DataFrame(dict_feature_count.items(), columns=['Feature', 'Count'])

    return(list_top10_df, df_feature_count)

##########################################################################################
#                                      Graph Functions                                   #
##########################################################################################

# Removes duplicates in legend
def legend_without_duplicate_labels(ax):
    handles, labels = ax.get_legend_handles_labels()
    unique = [(h, l) for i, (h, l) in enumerate(zip(handles, labels)) if l not in labels[:i]]
    ax.legend(*zip(*unique))


# Plots multiple Id - Vds curves on a single graph
def id_vds_plotter():
    df_original = prestatic_ds.copy()
    df_original.drop(df_original.columns[500:10026], axis=1,inplace=True)
    # df_original.drop(df_original[df_original['Device ID'].str.startswith('TR')].index, inplace=True)
    df_original = df_original.set_index('Device ID')
    # list_dev_id = df_original['Device ID'].tolist()
    # list_dev_id = df_original.index.values.tolist()

    # print(df_original)

    # Removes all of the devices that won't be plotted
    list_devices_to_plot = ['CREE1200V_1', 'LTF1200V_1', 'ROHM1700V_1', 'Infineon1200V_21', 'ST1200V_21']
    for dev_id, col in df_original.iterrows():
        if not list_devices_to_plot.count(dev_id) > 0:
            df_original = df_original.drop(dev_id)

    list_voltage = []
    dict_current = {}
    dict_passfail = df_pre_static_all.set_index('Device ID').to_dict()['Pass/Fail']

    # Strips everything but the number from column headings to get the Current value and creates a list
    for index, row in df_original.iteritems():
        num_voltage = re.sub('[^\d.-]', '', index)
        list_voltage.append(float(num_voltage))

    # Creates a dictionary, mapping Device ID to their voltage values
    for index, row in df_original.iterrows():
        dict_current[index] = row.values.tolist()

    # fig, ax = plt.subplots()
    fig = plt.figure(figsize = (8,8))
    ax = fig.add_subplot(1,1,1)

    # Iterates through dict_current and creates a graph
    for dev_id, list_current in dict_current.items():
        # Creates a dataframe for each device
        df_to_plot = pd.DataFrame()
        df_to_plot['Id(A)'] = list_current
        df_to_plot['Vds(V)'] = list_voltage

        # Gets whether or not the device passed or failed
        passfail = dict_passfail.get(dev_id, None)

        if passfail=='failed':
            ax.plot('Vds(V)', 'Id(A)', data=df_to_plot, marker='', color='red', linewidth=2, label='Failed')
        elif passfail=='good':
            ax.plot('Vds(V)', 'Id(A)', data=df_to_plot, marker='', color='green', linewidth=2, label='Passed')

    df_summary, df_summary_importance = model_iterator(df_pre_static_all, None, 'RFC', False)
    list_summary_top10dfs, df_summary_feature_count = find_top10(df_summary_importance)

    list_feature_count = df_summary_feature_count['Feature'].values.tolist()
    list_ds_feature_count = []

    # Gets the top 10 drain-source features
    for feature in list_feature_count:
        if len(list_ds_feature_count)==10:
            break
        elif (feature.startswith('I at')) and (not feature.endswith('_GS')):
            stripped_feature = float(re.sub('[^\d.-]', '', feature))
            list_ds_feature_count.append(stripped_feature)

    # Plots the top 10 drain-source features as vertical lines
    for xc in list_ds_feature_count:
        plt.axvline(x=xc, color='.1', linewidth=.3, alpha=.3, linestyle='--')

    # plt.legend()
    plt.title('Id - Vds')
    plt.xlabel('Vds(V)')
    plt.ylabel('Id(A)')

    # Remove duplicates in legend
    legend_without_duplicate_labels(ax)

    plt.show()

    df_original

# Visualizes data in a scatterplot with Principal Component Analysis
# x_values allows you to choose between using all of the features (All), only the drain-source features (DS), 
# and gate-source('GS')
# target_values allows you to specify how the entries are divded - Pass/Fail (PF) or Manufacturer (MN)
def pca_grapher(df_original: pd.DataFrame, x_values: str, target_values: str, drop_tr: bool, manu_to_keep: str, 
                output: bool, file_name: str):
    pd.set_option("display.max_rows", None, "display.max_columns", None)
    
    if drop_tr==True:
        # Creates a dataframe that drops TR650V and TR900V devices
        df = df_original.copy()
        df_dev_id = df.loc[:, ['Device ID']]
        df_dev_id.drop(df_dev_id[df_dev_id['Device ID'].str.startswith('TR')].index, inplace=True)
        df.drop(df[df['Device ID'].str.startswith('TR')].index, inplace=True)
    elif drop_tr==False:
        # Includes all of the devices
        df = df_original.copy()
        df_dev_id = df.loc[:, ['Device ID']]
        
    # Keeps all of the manufacturers
    if manu_to_keep=='ALL':
        pass
    # Drops all manufacturer but one. Make sure drop_tr is False when doing this
    else:
        df_dev_id.drop(df_dev_id[~df_dev_id['Device ID'].str.startswith(manu_to_keep)].index, inplace=True)
        df.drop(df[~df['Device ID'].str.startswith(manu_to_keep)].index, inplace=True)
    
    # Drops Infineon1200V_1 in both the main df and device id df
    df.drop(df[df['Device ID'] == 'Infineon1200V_1'].index, inplace=True)
    df.reset_index(drop=True,inplace=True)

    df_dev_id.drop(df_dev_id[df_dev_id['Device ID'] == 'Infineon1200V_1'].index, inplace=True)
    df_dev_id.reset_index(drop=True,inplace=True)
    
    # Includes all features and standardizes them
    if x_values=='ALL' or x_values==None:
        x = df.drop(['Device ID','Pass/Fail'], axis=1)
#         x = df.drop(['Pass/Fail'], axis=1)
        x = StandardScaler().fit_transform(x)
    # Drain-Source features
    elif x_values=='DS':
        x = df.drop(df.columns[501:1026], axis=1)
        x = x.drop('Device ID', axis=1)
    # Gate-Source features
    elif x_values=='GS':
        x = df.drop(df.columns[:501], axis=1)
        x = x.drop(df.columns[1001:1026], axis=1)
#         x = x.drop(['Temperature', 'Corrected Bias Voltage', 'Pass/Fail'], axis=1)

    # print(np.mean(x),np.std(x))
    fig = plt.figure(figsize = (3.5,3), dpi=600)
    pca = PCA(n_components=2)
    principalComponents = pca.fit_transform(x)
    # principalComponents = pca.fit(x)
    principalDf = pd.DataFrame(data = principalComponents, columns = ['principal component 1', 'principal component 2'])
    
    # Creates plot
    
    plt.rcParams['font.family'] = 'serif'
    plt.rcParams['font.serif'] = ['Times New Roman'] + plt.rcParams['font.serif']
    plt.rcParams['mathtext.default'] = 'regular'
    ax = fig.add_subplot(1,1,1) 
    ax.set_xlabel('Principal Component 1\n(b)', fontsize = 9)
    ax.set_ylabel('Principal Component 2', fontsize = 9)
    ax.set_title('2 Component PCA', fontsize = 10)    
    plt.xticks(size=8)
    plt.yticks(size=8)
#     ax.set_xscale('symlog')
#     ax.set_yscale('symlog')
#     ax.set_ylim(ymin=-150, ymax=200)
#     ax.set_xlim(xmin=-150, xmax=-80)
#     colors = np.random.rand(24,3)
    colors = ['r', 'g', 'b', 'c', 'm', 'y', 'k', 'tab:blue', 'tab:orange', 'tab:purple', 'tab:brown', 'tab:pink', 
              'tab:grey', 'tan', 'limegreen', 'springgreen', 'aqua', 'maroon', 'salmon', 'slateblue', 'cornflowerblue',
             'darkslateblue', 'mediumpurple', 'gold', 'darkseagreen']
    
    # Colors the scatterplot based on which target is chosen
    # Pass/Fail
    if target_values=='PF' or target_values==None:
        finalDf = pd.concat([principalDf, df_dev_id['Device ID'], df[['Pass/Fail']]], axis = 1)
        targets = ['good', 'failed']
#         colors = ['r', 'g', 'b', 'c', 'm', 'y', 'k', 'w']
        
        for target, color in zip(targets,colors):
            indicesToKeep = finalDf['Pass/Fail'] == target
            ax.scatter(finalDf.loc[indicesToKeep, 'principal component 1']
                       , finalDf.loc[indicesToKeep, 'principal component 2']
                       , c = color
                       , s = 15)
    # Manufacturer
    elif target_values=='MN':
        list_manu_dev = df_dev_id['Device ID'].to_list()
        list_manu = []

        for manu_dev in df_dev_id['Device ID']:
            manu = manu_dev.split('_')[0]
            list_manu.append(manu)

        df_dev_id['Device ID'].replace(list_manu_dev, list_manu, inplace=True)
        
        finalDf = pd.concat([principalDf, df_dev_id[['Device ID']]], axis = 1)
        targets = list(dict.fromkeys(list_manu))
        
        for target, color in zip(targets,colors):
            indicesToKeep = finalDf['Device ID'] == target
            ax.scatter(finalDf.loc[indicesToKeep, 'principal component 1']
                       , finalDf.loc[indicesToKeep, 'principal component 2']
                       , c = color
                       , s = 15)
    # Device Number
    elif target_values=='DN':
        list_manu_dev = df_dev_id['Device ID'].to_list()
        list_devnum = []

        for manu_dev in df['Device ID']:
            devnum = manu_dev.split('_')[1]
            list_devnum.append(devnum)
        
        df_dev_num = df_dev_id.copy()
        df_dev_num['Device ID'].replace(list_manu_dev, list_devnum, inplace=True)
        df_dev_num.rename(columns={'Device ID':'Device Number'}, inplace=True)
        
        finalDf = pd.concat([principalDf, df_dev_id['Device ID'], df_dev_num[['Device Number']]], axis = 1)
        targets = list(dict.fromkeys(list_devnum))
#         colors = np.random.rand(24,3)
        
        for target, color in zip(targets,colors):
            indicesToKeep = finalDf['Device Number'] == target
            ax.scatter(finalDf.loc[indicesToKeep, 'principal component 1']
                       , finalDf.loc[indicesToKeep, 'principal component 2']
                       , c = color
                       , s = 15)

    plt.legend(targets, prop={'size':8}, loc='best') #bbox_to_anchor=(1.05, 1.0)
    #ax.grid()
    fig.tight_layout()

    print('Explained variation ratio per principal component: {}'.format(pca.explained_variance_ratio_))
#     print(pca.explained_variance_)
    
    if output==True:
        plt.savefig(file_name+'.png', dpi=600, bbox_inches='tight')
    
    return finalDf

# Visualizes data in a scatterplot with Principal Component Analysis
# x_values allows you to choose between using all of the features (All), only the drain-source features (DS),
# and gate-source('GS')
# target_values allows you to specify how the entries are divded - Pass/Fail (PF) or Manufacturer (MN)
def pca_grapher_broken_axis(df_original: pd.DataFrame, x_values: str, target_values: str, drop_tr: bool,
                            manu_to_keep: str, output: bool, file_title: str):
    pd.set_option("display.max_rows", None, "display.max_columns", None)

    if drop_tr==True:
        # Creates a dataframe that drops TR650V and TR900V devices
        df = df_original.copy()
        df_dev_id = df.loc[:, ['Device ID']]
        df_dev_id.drop(df_dev_id[df_dev_id['Device ID'].str.startswith('TR')].index, inplace=True)
        df.drop(df[df['Device ID'].str.startswith('TR')].index, inplace=True)
    elif drop_tr==False:
        # Includes all of the devices
        df = df_original.copy()
        df_dev_id = df.loc[:, ['Device ID']]

    # Keeps all of the manufacturers
    if manu_to_keep=='ALL':
        pass
    # Drops all manufacturer but one. Make sure drop_tr is False when doing this
    else:
        df_dev_id.drop(df_dev_id[~df_dev_id['Device ID'].str.startswith(manu_to_keep)].index, inplace=True)
        df.drop(df[~df['Device ID'].str.startswith(manu_to_keep)].index, inplace=True)

    # Drops Infineon1200V_1 in both the main df and device id df
    df.drop(df[df['Device ID'] == 'Infineon1200V_1'].index, inplace=True)
    df.reset_index(drop=True,inplace=True)

    df_dev_id.drop(df_dev_id[df_dev_id['Device ID'] == 'Infineon1200V_1'].index, inplace=True)
    df_dev_id.reset_index(drop=True,inplace=True)

    # Includes all features and standardizes them
    if x_values=='ALL' or x_values==None:
        x = df.drop(['Device ID','Pass/Fail'], axis=1)
#         x = df.drop(['Pass/Fail'], axis=1)
        x = StandardScaler().fit_transform(x)
    # Drain-Source features
    elif x_values=='DS':
        x = df.drop(df.columns[501:1026], axis=1)
        x = x.drop('Device ID', axis=1)
    # Gate-Source features
    elif x_values=='GS':
        x = df.drop(df.columns[:501], axis=1)
        x = x.drop(df.columns[1001:1026], axis=1)
#         x = x.drop(['Temperature', 'Corrected Bias Voltage', 'Pass/Fail'], axis=1)

    # print(np.mean(x),np.std(x))

    pca = PCA(n_components=2)
    principalComponents = pca.fit_transform(x)
    # principalComponents = pca.fit(x)
    principalDf = pd.DataFrame(data = principalComponents, columns = ['principal component 1', 'principal component 2'])

#     # Creates plot
#     fig = plt.figure(figsize = (8,8))
#     ax = fig.add_subplot(2,1,1)
#     ax.set_xlabel('Principal Component 1', fontsize = 15)
#     ax.set_ylabel('Principal Component 2', fontsize = 15)
#     ax.set_title('2 component PCA', fontsize = 20)
#     ax.set_xscale('symlog')
#     ax.set_yscale('symlog')
#     ax.set_ylim(ymin=-150, ymax=200)
#     ax.set_xlim(xmin=-150, xmax=-80)

    # If we were to simply plot pts, we'd lose most of the interesting
    # details due to the outliers. So let's 'break' or 'cut-out' the y-axis
    # into two portions - use the top (ax1) for the outliers, and the bottom
    # (ax2) for the details of the majority of our data
    fig, (ax1, ax2) = plt.subplots(1, 2, sharey=True)
    fig.subplots_adjust(hspace=0.05)  # adjust space between axes
    plt.rcParams["figure.figsize"] = [30,15]
#     ax1.set_xscale('symlog')
#     ax1.set_yscale('symlog')
#     ax2.set_xscale('symlog')
#     ax2.set_yscale('symlog')

#     ax2.set_xlabel('Principal Component 1', fontsize = 25)
    ax1.set_ylabel('Principal Component 2', weight='bold', fontsize = 30)

#     ax2.set_title('2 component PCA', fontsize = 30)

#     colors = np.random.rand(24,3)
    colors = ['r', 'g', 'b', 'c', 'm', 'y', 'k', 'tab:blue', 'tab:orange', 'tab:purple', 'tab:brown', 'tab:pink',
              'tab:grey', 'tan', 'limegreen', 'springgreen', 'aqua', 'maroon', 'salmon', 'slateblue', 'cornflowerblue',
             'darkslateblue', 'mediumpurple', 'gold', 'darkseagreen']

    # Colors the scatterplot based on which target is chosen
    # Pass/Fail
    if target_values=='PF' or target_values==None:
        finalDf = pd.concat([principalDf, df_dev_id['Device ID'], df[['Pass/Fail']]], axis = 1)
        targets = ['good', 'failed']
#         colors = ['r', 'g', 'b', 'c', 'm', 'y', 'k', 'w']

        for target, color in zip(targets,colors):
            indicesToKeep = finalDf['Pass/Fail'] == target
            ax1.scatter(finalDf.loc[indicesToKeep, 'principal component 1']
                   , finalDf.loc[indicesToKeep, 'principal component 2']
                   , c = color
                   , s = 15)

            ax2.scatter(finalDf.loc[indicesToKeep, 'principal component 1']
                   , finalDf.loc[indicesToKeep, 'principal component 2']
                   , c = color
                   , s = 15)
    # Manufacturer
    elif target_values=='MN':
        list_manu_dev = df_dev_id['Device ID'].to_list()
        list_manu = []

        for manu_dev in df_dev_id['Device ID']:
            manu = manu_dev.split('_')[0]
            list_manu.append(manu)

        df_dev_id['Device ID'].replace(list_manu_dev, list_manu, inplace=True)

        finalDf = pd.concat([principalDf, df_dev_id[['Device ID']]], axis = 1)
        targets = list(dict.fromkeys(list_manu))

        for target, color in zip(targets,colors):
            indicesToKeep = finalDf['Device ID'] == target

            ax1.scatter(finalDf.loc[indicesToKeep, 'principal component 1']
                   , finalDf.loc[indicesToKeep, 'principal component 2']
                   , c = color
                   , s = 15)

            ax2.scatter(finalDf.loc[indicesToKeep, 'principal component 1']
                   , finalDf.loc[indicesToKeep, 'principal component 2']
                   , c = color
                   , s = 15)

    # Device Number
    elif target_values=='DN':
        list_manu_dev = df_dev_id['Device ID'].to_list()
        list_devnum = []

        for manu_dev in df['Device ID']:
            devnum = manu_dev.split('_')[1]
            list_devnum.append(devnum)

        df_dev_num = df_dev_id.copy()
        df_dev_num['Device ID'].replace(list_manu_dev, list_devnum, inplace=True)
        df_dev_num.rename(columns={'Device ID':'Device Number'}, inplace=True)

        finalDf = pd.concat([principalDf, df_dev_id['Device ID'], df_dev_num[['Device Number']]], axis = 1)
        targets = list(dict.fromkeys(list_devnum))
#         colors = np.random.rand(24,3)

        for target, color in zip(targets,colors):
            indicesToKeep = finalDf['Device Number'] == target
            ax.scatter(finalDf.loc[indicesToKeep, 'principal component 1']
                       , finalDf.loc[indicesToKeep, 'principal component 2']
                       , c = color
                       , s = 15)
    if x_values=='DS':
        ax1.set_xlim(-150,-100)
        ax2.set_xlim(420,500)
    elif x_values=='ALL':
        ax1.set_xlim(-30,5)
        ax2.set_xlim(30,40)

    # For when All manu, All features
#     ax1.set_xlim(-30, 5)
#     ax2.set_xlim(30, 40)

    # hide the spines between ax and ax2
    ax1.spines['right'].set_visible(False)
    ax2.spines['left'].set_visible(False)
    ax1.yaxis.tick_left()
#     ax1.tick_params(labelright='off')
    ax2.yaxis.tick_right()
    ax2.tick_params(labelright='off')

#     # This looks pretty good, and was fairly painless, but you can get that
#     # cut-out diagonal lines look with just a bit more work. The important
#     # thing to know here is that in axes coordinates, which are always
#     # between 0-1, spine endpoints are at these locations (0,0), (0,1),
#     # (1,0), and (1,1).  Thus, we just need to put the diagonals in the
#     # appropriate corners of each of our axes, and so long as we use the
#     # right transform and disable clipping.

#     d = .015 # how big to make the diagonal lines in axes coordinates
#     # arguments to pass plot, just so we don't keep repeating them
#     kwargs = dict(transform=ax1.transAxes, color='k', clip_on=False, label=None)
#     ax1.plot((1-d,1+d), (-d,+d), **kwargs)
#     ax1.plot((1-d,1+d),(1-d,1+d), **kwargs)

#     kwargs.update(transform=ax2.transAxes, label=None)  # switch to the bottom axes
#     ax2.plot((-d,+d), (1-d,1+d), **kwargs)
#     ax2.plot((-d,+d), (-d,+d), **kwargs)

    ax1.tick_params(labelsize=25)
    ax2.tick_params(labelsize=25)

    plt.figtext(.5, .05, 'Principal Component 1', fontsize=30, weight='bold', ha='center', va='bottom')
    plt.figtext(.5, .95, '2-Component PCA', fontsize=35, weight='bold', ha='center', va='top')

    ax1.legend(targets, loc='upper right', bbox_to_anchor=(2.8, 1), prop={'size': 25})
#     ax2.legend(targets, loc='upper right', prop={'size': 25})
#     ax2.grid()
#     print(finalDf)

    print('Explained variation ratio per principal component: {}'.format(pca.explained_variance_ratio_))
#     print(pca.explained_variance_)

    if output==True:
        plt.savefig(base_path+'Plots/PassFail/'+file_title+'.png', dpi=150, bbox_inches='tight',pad_inches=1)

    return finalDf

# Creates a dendrogram depending on the specified input
# Use 'lastp' to truncate
def dendrogram(df_main: pd.DataFrame, features: str, clusters: int, truncate: bool, output_figure: bool,
               figure_title: str, file_title: str):
    # Across all features
    if features=='ALL':
        df_hc = df_main.copy()
        # df_hc = df_hc.drop(['Device ID', 'Pass/Fail'], axis=1)
        df_hc = df_hc.drop(['Pass/Fail'], axis=1)
        df_hc = df_hc.set_index('Device ID')
    # Drain-Source features only
    elif features=='DS':
        df_hc = df_main[df_main.columns[:501]]
        df_hc = df_hc.set_index('Device ID')
#         df_hc = df_hc.drop(['Device ID'], axis=1)
    # Gate-Source features only
    elif features=='GS':
        df_hc = df_main.copy()
#         df_hc.drop(df_hc[df_hc.columns[501:1025]]
        df_hc = df_hc.set_index('Device ID')
        df_hc.drop(df_hc.columns[~df_hc.columns.str.endswith('GS')], axis=1, inplace=True)
    # Everything but Gate-Source features
    elif features=='No GS':
        df_hc = df_main.copy()
        df_hc = df_hc.set_index('Device ID')
        df_hc = df_hc.drop(['Pass/Fail'], axis=1)
        df_hc.drop(df_hc.columns[df_hc.columns.str.endswith('GS')], axis=1, inplace=True)

    # Create a dendrogram
    plt.figure(figsize=(30, 15))
    plt.title(figure_title, fontsize=35, weight='bold')
    plt.xlabel('Devices', fontsize=30, weight='bold')
    plt.ylabel('Euclidean distances', fontsize=30, weight='bold')
    plt.xticks(size=25)
    plt.yticks(size=25)
#     plt.xticks(fontsize=200)
#     plt.tick_params(axis='both', which='minor', labelsize=200)
#     plt.rc('xtick',labelsize=200)

    linkage_matrix = shc.linkage(df_hc, method='ward')

    if truncate==True:
#         list_labels = random.sample(range(300), 222)
        list_labels = df_hc.index.values.tolist()
        dendrogram = shc.dendrogram(linkage_matrix, p=3, labels=list_labels, truncate_mode='level',
                                    show_leaf_counts=True, show_contracted=False)
    elif truncate==False:
        dendrogram = shc.dendrogram(linkage_matrix, labels=df_hc.index.values.tolist(), truncate_mode=None,
                                    show_leaf_counts=True, show_contracted=False)
#     plt.axhline(y=4300, color='r', linestyle='--')

    cluster = AgglomerativeClustering(n_clusters=clusters, affinity='euclidean', linkage='ward')
    cluster.fit_predict(df_hc)

    # Converts the dendrogram into a png
    if output_figure==True:
        plt.savefig(base_path+'Plots/PassFail/'+file_title+'.png', dpi=150)

# Creates a specific truncated dendrogram with labels
def manufacturer_dendrogram(label_type: str, output: bool):
    df_hc = df_pre_static[df_pre_static.columns[:501]]
    df_hc = df_hc.set_index('Device ID')

    linked = shc.linkage(df_hc, 'ward')

    if label_type=='full_name':
        labels = ['', '', 'I', '&', 'J', '', '', '', 'E', 'G', 'B',
                  'B', 'C', 'B & D', 'F', 'H']
        file_title = 'HC_TruncatedDendro_Manu_Fullnames'
    elif label_type=='codename':
        labels = ['', '', 'H650V', '&', 'I900V', '', '', '', 'E1700V', 'J1700V', 'G1700V', 'G1700V', 'F1200V',
                  'G1700V', 'G1700V & D1200V', 'B1200V', 'A1200V']
        file_title = 'HC_TruncatedDendro_Manu_Codenames'

    p = len(labels)

    plt.figure(figsize=(3.5,3), dpi=600)
    plt.title('HC Dendrogram - Manufacturer Clusters', fontsize=10)
    plt.xlabel('Distance\n(a)', fontsize=9)
    plt.ylabel('Manufacturers', fontsize=9)
    plt.xticks(size=8)
    plt.yticks(size=8)

    # call dendrogram to get the returned dictionary
    # (plotting parameters can be ignored at this point)
    R = shc.dendrogram(
                    linked,
                    truncate_mode='lastp',  # show only the last p merged clusters
                    p=p,  # show only the last p merged clusters
                    no_plot=True,
                    )

    print('values passed to leaf_label_func\nleaves : ', R['leaves'])

    # create a label dictionary
    dict_labels = {R['leaves'][ii]: labels[ii] for ii in range(len(R['leaves']))}
    def llf(xx):
        return '{}'.format(dict_labels[xx])

    ## This version gives you your label AND the count
    # dict_labels = {R['leaves'][ii]:(labels[ii], R['ivl'][ii]) for ii in range(len(R['leaves']))}
    # def llf(xx):
    #     return '{} - {}'.format(*dict_labels[xx])

    shc.dendrogram(
                linked,
                orientation='right',
                truncate_mode='lastp',  # show only the last p merged clusters
                p=p,  # show only the last p merged clusters
                leaf_label_func=llf,
                leaf_rotation=0,
                leaf_font_size=9,
                show_contracted=False,  # to get a distribution impression in truncated branches
                )
#     plt.show()

    if output==True:
        plt.savefig(file_title+'.png', dpi=600, bbox_inches='tight')

# Box plot with model iterator divided by manufacturer

def boxplotter(df: pd.DataFrame, x_val: str, y_val: str, x_label: str, y_label: str, title: str,
               file_title: str, output: bool):
    plt.rcParams['font.family'] = 'serif'
    plt.rcParams['font.serif'] = ['Times New Roman'] + plt.rcParams['font.serif']
    plt.rcParams['mathtext.default'] = 'regular'
    fig = plt.figure(figsize = (7.16,3.5), dpi=600)
#     ax = fig.add_subplot(1,1,1)

    ax = sns.boxplot(x=x_val, y=y_val, data=df)
    plt.xticks(rotation=30, size=5)
    plt.yticks(size=8)
    plt.xlabel(x_label, fontsize=9)
    plt.ylabel(y_label, fontsize=9)
    plt.title(title, fontsize=10)

    if output==True:
        plt.savefig(file_title+'.pdf', dpi=600, bbox_inches='tight')

##########################################################################################
#                              Miscellaneous Functions                                   #
##########################################################################################

# Adds dataframe used in this code to the device_pipeline.xlsx
def add_df_to_overview():
    new_df = df_pre_static_all.copy()
    new_df['Pass/Fail'] = new_df['Pass/Fail'].replace(['failed', 'good'], ['Fail', 'Pass'])

    sheet_creator(base_path+'device_pipeline.xlsx', 'PassFail_Features_Dataframe', new_df, False)

    return new_df
# add_pf_to_overview()

# Adds the material that transistors are made of (TR - Gallium Nitride, Everything else - Silicon Carbide) to
# device_pipeline.xlsx to a new column
def material_appender(df: pd.DataFrame, excel_name: str, sheet_name: str):
    dict_mat = {}
    df_new = df.copy()

    for dev_id in df['Device ID']:
        if dev_id.startswith('TR'):
            dict_mat[dev_id] = 'GaN'
        else:
            dict_mat[dev_id] = 'SiC'

    df_new['Material'] = df_new['Device ID'].map(dict_mat)

    sheet_creator(excel_name, sheet_name, df_new)

# material_appender(overview, base_path+'device_pipeline.xlsx', 'Dynamic Devices Overview')

# Adds multiple dataframes to one excel sheet
# Specifically for the Top10Features sheet
def multiple_dfs(df_list, under_df: bool, df_count, sheets, file_name, spaces):
    book = load_workbook(file_name)
    writer = pd.ExcelWriter(file_name, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    col = 0
    for dataframe in df_list:
        dataframe.to_excel(writer,sheet_name=sheets,startrow=0,startcol=col, index=False)
#         col = col + spaces + 2
        col = len(dataframe.columns) + spaces

    if under_df==True:
        df_count.to_excel(writer,sheet_name=sheets,startrow=12,startcol=0,index=False)

    writer.save()

# multiple_dfs(list_top10_df, 'Top10Features', base_path+'passfail_output.xlsx', 1)

# Adds df_model_acc to device_pipeline
def sheet_creator(excel_path: str, excel_sheet_name: str, df: pd.DataFrame, ind: bool):

    if os.path.exists(excel_path):

        # Path for device_pipeline
        book = load_workbook(excel_path)
        writer = pd.ExcelWriter(excel_path, engine='openpyxl')
        writer.book = book

        ## ExcelWriter for some reason uses writer.sheets to access the sheet.
        ## If you leave it empty it will not know that sheet Main is already there
        ## and will create a new sheet.

        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        df.to_excel(writer, sheet_name=excel_sheet_name, index=ind)

        writer.save()

    else:
        df.to_excel(excel_path, sheet_name=excel_sheet_name)


# Function that performs one-hot encoding
def encode_and_bind(original_dataframe, feature_to_encode):
    dummies = pd.get_dummies(original_dataframe[[feature_to_encode]])
    res = pd.concat([original_dataframe, dummies], axis=1)
    return(res)

# One-Hot Encoder that adds Manufactuerer+Voltage (CREE1200V, etc.), Manufacturer (CREE, etc.), Voltage (1200V, etc.)
# and Material (SiC or GaN) to device_pipeline.xlsx in a sheet titled 'One Hot Encoded'
def one_hot_encoder():
    # Creates a dataframe that consists of Device ID's
    df_device_id_separated = pd.DataFrame(df_pre_static_all['Device ID'])

    # Creates dictionaries to be added to the dataframe
    dict_manu = {}
    dict_volt = {}
    dict_manu_volt = {}
    dict_material = overview.set_index('Device ID').to_dict()['Material']

    # Gets manufacturers+voltages, manufacturers, and voltages and adds them to their respective dictionaries
    for device_id in df_pre_static_all['Device ID']:
        manu_volt = device_id.split('_')[0]
        manu = re.sub('[\dV]','', manu_volt)
        volt = re.sub(manu, '', manu_volt)
        dict_manu_volt[device_id] = manu_volt
        dict_manu[device_id] = manu
        dict_volt[device_id] = volt

    # Adds dictionaries to dataframes as columns
    df_device_id_separated['Manufacturer and Voltage'] = df_device_id_separated['Device ID'].map(dict_manu_volt)
    df_device_id_separated['Manufacturer'] = df_device_id_separated['Device ID'].map(dict_manu)
    df_device_id_separated['Voltage'] = df_device_id_separated['Device ID'].map(dict_volt)
    df_device_id_separated['Material'] = df_device_id_separated['Device ID'].map(dict_material)

    # Creates a list of the columns that will be encoded
    list_features_to_encode = ['Manufacturer and Voltage', 'Manufacturer', 'Voltage', 'Material']

    # One-hot encodes the items and adds them to the dataframe
    for i in list_features_to_encode:
        df_device_id_separated = encode_and_bind(df_device_id_separated, i)

    # Exports df_device_id_separated to device_pipleline.xlsx
    # sheet_creator(base_path+'device_pipeline.xlsx', 'One Hot Encoded', df_device_id_separated, False)




